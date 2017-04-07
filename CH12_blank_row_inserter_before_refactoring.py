#! python3
import openpyxl
import sys
import os

while True:
    os.chdir('C:\\Python')
    if len(sys.argv) < 4:
        print('Not enough data! [row] [blank] [file name]')
        break
    else:
        rowNum = int(sys.argv[1])
        blankNum = int(sys.argv[2])
        fileName = sys.argv[3]
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.active
    if sheet.max_row <= rowNum:
        print('Document to small to make a blank row in this place!')
        break
    else:
        newWb = openpyxl.Workbook()
        newSheet = newWb.active
        for tupleOfRows in sheet.rows[:rowNum]:
            for cellObj in tupleOfRows:
                newSheet[cellObj.column+str(cellObj.row)].value = cellObj.value
        for tupleOfRows in sheet.rows[rowNum:]:
            for cellObj in tupleOfRows:
                newSheet[cellObj.column+str(cellObj.row+blankNum)].value = cellObj.value
        newWb.save('blank'+fileName)
        break
