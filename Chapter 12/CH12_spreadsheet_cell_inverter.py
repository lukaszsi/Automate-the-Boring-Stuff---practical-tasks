#! python3
import openpyxl
import os

from openpyxl.cell import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
newWb = openpyxl.Workbook()
newSheet = newWb.active
for tupleOfRows in sheet.rows:
    for cellObj in tupleOfRows:
        newSheet.cell(
            row=column_index_from_string(cellObj.column), 
            column=cellObj.row
        ).value = cellObj.value
newWb.save('inverted_example.xlsx')
