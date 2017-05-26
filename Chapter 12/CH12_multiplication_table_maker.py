#! python3
import sys, os, openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font

if len(sys.argv) < 2:
    print(' I need a number!')
else:
    arg = sys.argv[1]
    
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1,int(arg)+1):
    sheet.cell(row=1, column=i+1).value = i
    sheet.cell(row=i+1, column=1).value = i
for rowOfCellObjects in sheet['B2':get_column_letter(int(arg)+1) + str(int(arg) + 1)]:
    for cellObj in rowOfCellObjects:
        cellObj.value = sheet.cell(
            row=cellObj.row, column=1).value * sheet.cell(
            row=1, column=column_index_from_string(cellObj.column)).value
boldFont = Font(bold=True)
for cellObj in sheet.columns[0]:
    cellObj.font = boldFont
for cellObj in sheet.rows[0]:
    cellObj.font = boldFont
os.chdir('c:/Python')
wb.save('multiplicationTable.xlsx')
